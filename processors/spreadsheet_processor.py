"""
SpreadsheetProcessor
====================
A single-file, production-ready-looking Python module to extract exhaustive metadata
and content from spreadsheet files (XLSX/XLSM/XLSB/XLS/ODS/CSV).

Notes:
- This module focuses on deterministic extraction (no decisions/judgement) and emits a
  comprehensive JSON-friendly dict described by the user.
- Dependencies (install via pip):
    pip install openpyxl olefile pandas python-docx pyxlsb odfpy python-magic
  (Some libraries are optional; the module falls back gracefully when they are missing.)

Limitations:
- No real system or ML decision-making is performed here; this module only extracts features.
- "0% false positives/negatives" is impossible to guarantee in the general world; this
  implementation strives for deterministic extraction but accuracy depends on the input and
  installed parser backends.

Usage example:
    from SpreadsheetProcessor import SpreadsheetProcessor
    with open('sample.xlsx','rb') as f:
        data = f.read()
    report = SpreadsheetProcessor.process(data, '.xlsx')

"""

from __future__ import annotations

import io
import tempfile
import zipfile
import csv
import re
import os
import hashlib
import json
from typing import Any, Dict, List, Tuple, Optional
from collections import Counter, defaultdict
from datetime import datetime

# Try optional imports
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
except Exception:
    openpyxl = None

try:
    import xlrd
except Exception:
    xlrd = None

try:
    from pyxlsb import open_workbook as open_xlsb
except Exception:
    open_xlsb = None

try:
    from odf.opendocument import load as load_ods
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
except Exception:
    load_ods = None

try:
    import olefile
except Exception:
    olefile = None

# Basic helpers

FUNCTION_REGEX = re.compile(r"([A-Z_][A-Z0-9_]*)\s*\(", re.IGNORECASE)
CELL_REF_REGEX = re.compile(r"\$?[A-Z]{1,3}\$?\d+")


def _sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _safe_decode_bytes(b: bytes) -> str:
    try:
        return b.decode('utf-8', errors='replace')
    except Exception:
        return str(b)


class SpreadsheetProcessor:
    """Main class exposing `process` method.

    process(cls, file_data: bytes, file_extension: str) -> Dict[str, Any]
    """

    @classmethod
    def process(cls, file_data: bytes, file_extension: str) -> Dict[str, Any]:
        """Entrypoint. Detects the type from extension and delegates.

        file_extension should be like '.xlsx', '.xlsm', '.xls', '.xlsb', '.ods', '.csv'
        """
        ext = (file_extension or '').lower()
        result: Dict[str, Any] = {
            'file_level_metadata': {},
            'sheets': [],
            'macros_and_scripts': {},
            'embedded_items': {},
            'external_references': {},
            'text_extraction': {},
            'numeric_and_date_analysis': {},
            'structural_features_for_ml': {},
        }

        # common hashes

        # Save to temp file for libraries that require a filename
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext if ext else '')
        try:
            tmp.write(file_data)
            tmp.flush()
            tmp.close()

            if ext in ('.xlsx', '.xlsm') and zipfile.is_zipfile(tmp.name) and openpyxl:
                cls._process_xlsx(tmp.name, file_data, result)
            elif ext == '.xlsb' and open_xlsb:
                cls._process_xlsb(tmp.name, file_data, result)
            elif ext == '.xls' and xlrd:
                cls._process_xls(tmp.name, file_data, result)
            elif ext == '.ods' and load_ods:
                cls._process_ods(tmp.name, file_data, result)
            elif ext == '.csv':
                cls._process_csv(io.BytesIO(file_data), result)
            else:
                # fallback: try to inspect as zip (xlsx/xlsm) or as plain csv
                if zipfile.is_zipfile(tmp.name) and openpyxl:
                    cls._process_xlsx(tmp.name, file_data, result)
                else:
                    try:
                        cls._process_csv(io.BytesIO(file_data), result)
                    except Exception:
                        # last resort: try to parse basic text
                        result['text_extraction']['all_text'] = _safe_decode_bytes(file_data)[:20000]
        finally:
            try:
                os.unlink(tmp.name)
            except Exception:
                pass

        # post-processing: structural features
        cls._finalize_structural_features(result)

        return result

    # ---------- parsers ----------
    @classmethod
    def _process_xlsx(cls, filename: str, file_data: bytes, result: Dict[str, Any]):
        # Use zipfile to inspect relationships, embedded objects, vba presence
        z = zipfile.ZipFile(filename, 'r')
        namelist = z.namelist()

        wb_has_vba = any(n.lower().endswith('vbaProject.bin'.lower()) for n in namelist)
        embedded = [n for n in namelist if n.startswith('xl/embeddings/')]
        charts = [n for n in namelist if n.startswith('xl/charts/')]
        external_links = [n for n in namelist if 'externalLinks' in n]

        result['file_level_metadata'].update({
            'has_vba': wb_has_vba,
            'has_embedded_objects': len(embedded) > 0,
            'num_charts': len(charts),
            'has_external_links': len(external_links) > 0,
        })

        # sheets & content using openpyxl
        try:
            wb = load_workbook(filename, data_only=False, read_only=True)
            sheet_names = wb.sheetnames
            result['file_level_metadata']['num_sheets'] = len(sheet_names)
            result['file_level_metadata']['sheet_names'] = sheet_names

            hidden_count = 0
            total_named_ranges = []
            for name in wb.defined_names.definedName:
                total_named_ranges.append(name.name)
            result['file_level_metadata']['named_ranges'] = total_named_ranges

            for ws in wb.worksheets:
                sdict = {
                    'sheet_name': ws.title,
                    'visible': not ws.sheet_state == 'hidden',
                }
                if not sdict['visible']:
                    hidden_count += 1

                # iterate rows to collect stats (read-only safe)
                row_count = 0
                col_count = 0
                non_blank = 0
                numeric_cells = 0
                text_cells = 0
                formula_cells = 0
                text_lengths = []
                total_cells = 0
                samples_top_left = []
                random_samples = []
                formula_funcs = Counter()

                max_rows_to_scan = 2000  # cap for performance while still giving good stats
                for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                    if r_idx > max_rows_to_scan:
                        break
                    row_count = max(row_count, r_idx)
                    for c_idx, cell in enumerate(row, start=1):
                        total_cells += 1
                        col_count = max(col_count, c_idx)
                        v = None
                        try:
                            v = cell.value
                        except Exception:
                            v = None
                        if v is None or (isinstance(v, str) and v.strip() == ''):
                            continue
                        non_blank += 1
                        if isinstance(v, (int, float)):
                            numeric_cells += 1
                        elif isinstance(v, str):
                            text_cells += 1
                            text_lengths.append(len(v))
                        # formula
                        if getattr(cell, 'data_type', None) == 'f' or (isinstance(v, str) and v.startswith('=')):
                            formula_cells += 1
                            ftext = v if isinstance(v, str) else getattr(cell, 'formula', '')
                            # extract functions
                            for fn in FUNCTION_REGEX.findall(str(ftext)):
                                formula_funcs[fn.upper()] += 1
                            # sample suspicious pattern
                            if any(keyword in str(ftext).lower() for keyword in ['encode', 'base64', 'shell', 'url', 'hyperlink', 'weblogin']):
                                random_samples.append({'r': f'{get_column_letter(c_idx)}{r_idx}', 'v': str(ftext), 'type': 'formula'})

                # top-left sample
                try:
                    top_left = []
                    for r in range(1, min(3, ws.max_row) + 1):
                        rowvals = []
                        for c in range(1, min(6, ws.max_column) + 1):
                            cv = ws.cell(row=r, column=c).value
                            rowvals.append(cv)
                        top_left.append(rowvals)
                    samples_top_left = top_left
                except Exception:
                    samples_top_left = []

                sdict['row_count'] = row_count
                sdict['col_count'] = col_count
                sdict['samples'] = {
                    'top_left_cells': samples_top_left,
                    'random_cell_samples': random_samples[:10]
                }
                sdict['cell_stats'] = {
                    'total_cells': total_cells,
                    'non_blank_cells': non_blank,
                    'numeric_cells': numeric_cells,
                    'text_cells': text_cells,
                    'formula_cells': formula_cells,
                    'avg_text_length': (sum(text_lengths) / len(text_lengths)) if text_lengths else 0,
                    'max_text_length': (max(text_lengths) if text_lengths else 0)
                }
                sdict['formulas'] = {
                    'count': formula_cells,
                    'top_functions': dict(formula_funcs.most_common(20)),
                }

                result['sheets'].append(sdict)

            # hidden sheets count
            result['file_level_metadata']['has_hidden_sheets'] = hidden_count > 0
            result['file_level_metadata']['num_hidden_sheets'] = hidden_count

            # embedded objects extraction
            emb_list = []
            for emb in embedded:
                try:
                    b = z.read(emb)
                    emb_list.append({'path': emb, 'sha256': hashlib.sha256(b).hexdigest()})
                except Exception:
                    emb_list.append({'path': emb, 'sha256': None})
            result['embedded_items']['excel_embedded_files'] = emb_list

            # external references and linked domains
            linked_domains = set()
            extlinks = []
            for name in namelist:
                if name.startswith('xl/externalLinks'):
                    try:
                        txt = z.read(name).decode('utf-8', errors='replace')
                        for m in re.findall(r'https?://[^\"\>\s<]+', txt):
                            linked_domains.add(re.sub(r'^https?://', '', m).split('/')[0])
                            extlinks.append(m)
                    except Exception:
                        pass
            result['external_references']['external_links'] = extlinks
            result['external_references']['linked_domains'] = list(linked_domains)

            # macros & vba
            if wb_has_vba:
                try:
                    vba_blob = z.read('xl/vbaProject.bin')
                    result['macros_and_scripts']['has_vba'] = True
                    result['macros_and_scripts']['vba_sha256'] = hashlib.sha256(vba_blob).hexdigest()
                    # quick heuristic: search text for obvious function names like shell or createobject
                    s = re.sub(r'[^\x00-\x7f]', '', vba_blob.decode('latin-1', errors='ignore'))
                    suspicious = []
                    for keyword in ['shell(', 'createobject', 'base64', 'shellexecute', 'urldownloadtofile', 'xmlhttp']:
                        if keyword in s.lower():
                            suspicious.append(keyword)
                    result['macros_and_scripts']['suspicious_indicators'] = list(set(suspicious))
                except Exception:
                    pass
            else:
                result['macros_and_scripts']['has_vba'] = False

            # text extraction: collect all text-like parts from workbook for ML
            all_text_parts = []
            # core props
            try:
                if 'docProps/core.xml' in namelist:
                    all_text_parts.append(z.read('docProps/core.xml').decode('utf-8', errors='replace'))
            except Exception:
                pass
            # sheet xmls
            for n in namelist:
                if n.startswith('xl/worksheets/'):
                    try:
                        t = z.read(n).decode('utf-8', errors='replace')
                        all_text_parts.append(t)
                    except Exception:
                        pass
            joined_text = '\n'.join(all_text_parts)[:200000]
            # simple PII/email extraction
            emails = re.findall(r'([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)', joined_text)
            result['text_extraction']['all_text'] = joined_text[:20000]
            result['text_extraction']['pii_counts'] = {'email': len(emails)}
            result['text_extraction']['languages_detected'] = ['en'] if re.search(r'[a-zA-Z]', joined_text) else []

        except Exception as e:
            result['file_level_metadata']['parsing_error'] = str(e)

        finally:
            try:
                z.close()
            except Exception:
                pass

    @classmethod
    def _process_xlsb(cls, filename: str, file_data: bytes, result: Dict[str, Any]):
        # limited support using pyxlsb for cell-level read
        if open_xlsb is None:
            result['file_level_metadata']['parsing_error'] = 'pyxlsb library is not installed.'
            return
        try:
            sheets = []
            with open_xlsb(filename) as wb:
                for sheetname in wb.sheets:
                    row_count = 0
                    col_count = 0
                    total_cells = 0
                    non_blank = 0
                    numeric_cells = 0
                    text_cells = 0
                    formula_cells = 0
                    samples = []
                    for row in wb.get_sheet(sheetname):
                        row_count += 1
                        for c_idx, v in enumerate(row):
                            total_cells += 1
                            if v.v is None:
                                continue
                            non_blank += 1
                            if isinstance(v.v, (int, float)):
                                numeric_cells += 1
                            elif isinstance(v.v, str):
                                text_cells += 1
                            # formulas in xlsb not easily available here
                    sheets.append({'sheet_name': sheetname, 'row_count': row_count, 'total_cells': total_cells})
            result['file_level_metadata']['num_sheets'] = len(sheets)
            result['sheets'] = sheets
        except Exception as e:
            result['file_level_metadata']['parsing_error'] = str(e)

    @classmethod
    def _process_xls(cls, filename: str, file_data: bytes, result: Dict[str, Any]):
        # Older BIFF .xls using xlrd (if available)
        if xlrd is None:
            result['file_level_metadata']['parsing_error'] = 'xlrd library is not installed.'
            return
        try:
            wb = xlrd.open_workbook(filename)
            sheet_names = wb.sheet_names()
            result['file_level_metadata']['num_sheets'] = len(sheet_names)
            result['file_level_metadata']['sheet_names'] = sheet_names
            for name in sheet_names:
                sh = wb.sheet_by_name(name)
                total_cells = sh.nrows * sh.ncols
                non_blank = 0
                numeric = 0
                text = 0
                formulas = 0
                top_left = []
                for r in range(min(3, sh.nrows)):
                    rowvals = []
                    for c in range(min(6, sh.ncols)):
                        try:
                            v = sh.cell_value(r, c)
                        except Exception:
                            v = None
                        rowvals.append(v)
                        if v not in ('', None):
                            non_blank += 1
                            if isinstance(v, (int, float)):
                                numeric += 1
                            elif isinstance(v, str):
                                text += 1
                    top_left.append(rowvals)
                result['sheets'].append({'sheet_name': name, 'row_count': sh.nrows, 'col_count': sh.ncols, 'samples': {'top_left_cells': top_left}, 'cell_stats': {'total_cells': total_cells, 'non_blank_cells': non_blank, 'numeric_cells': numeric, 'text_cells': text}})
        except Exception as e:
            result['file_level_metadata']['parsing_error'] = str(e)

    @classmethod
    def _process_ods(cls, filename: str, file_data: bytes, result: Dict[str, Any]):
        if load_ods is None:
            result['file_level_metadata']['parsing_error'] = 'odfpy library is not installed.'
            return
        try:
            doc = load_ods(filename)
            tables = doc.getElementsByType(Table)
            result['file_level_metadata']['num_sheets'] = len(tables)
            for t in tables:
                name = t.getAttribute('name')
                rows = t.getElementsByType(TableRow)
                rowcount = len(rows)
                # simple sample
                top_left = []
                for r in range(min(3, rowcount)):
                    cells = rows[r].getElementsByType(TableCell)
                    rowvals = []
                    for c in range(min(6, len(cells))):
                        ps = cells[c].getElementsByType(P)
                        txt = ''.join([p.firstChild.data if p.firstChild is not None else '' for p in ps]) if ps else ''
                        rowvals.append(txt)
                    top_left.append(rowvals)
                result['sheets'].append({'sheet_name': name, 'row_count': rowcount, 'samples': {'top_left_cells': top_left}})
        except Exception as e:
            result['file_level_metadata']['parsing_error'] = str(e)

    @classmethod
    def _process_csv(cls, fileobj: io.BytesIO, result: Dict[str, Any]):
        try:
            fileobj.seek(0)
            decoded = io.TextIOWrapper(fileobj, encoding='utf-8', errors='replace')
            reader = csv.reader(decoded)
            rows = list(reader)
            rcount = len(rows)
            ccount = max((len(r) for r in rows), default=0)
            samples = rows[:5]
            total_cells = sum(len(r) for r in rows)
            non_blank = sum(1 for r in rows for c in r if c and str(c).strip())
            numeric = sum(1 for r in rows for c in r if cls._is_number(c))
            text_cells = non_blank - numeric
            result['file_level_metadata']['num_sheets'] = 1
            result['file_level_metadata']['sheet_names'] = ['CSV']
            result['sheets'].append({'sheet_name': 'CSV', 'row_count': rcount, 'col_count': ccount, 'samples': {'top_left_cells': samples}, 'cell_stats': {'total_cells': total_cells, 'non_blank_cells': non_blank, 'numeric_cells': numeric, 'text_cells': text_cells}})
            # text extraction
            joined = '\n'.join([','.join(r) for r in rows])
            emails = re.findall(r'([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)', joined)
            result['text_extraction']['all_text'] = joined[:20000]
            result['text_extraction']['pii_counts'] = {'email': len(emails)}
        except Exception as e:
            result['file_level_metadata']['parsing_error'] = str(e)

    # ---------- utilities ----------
    @staticmethod
    def _is_number(v: Any) -> bool:
        try:
            float(v)
            return True
        except Exception:
            return False

    @classmethod
    def _finalize_structural_features(cls, result: Dict[str, Any]):
        # Build features vector
        flm = result.get('file_level_metadata', {})
        sheets = result.get('sheets', [])
        num_sheets = flm.get('num_sheets', len(sheets))
        num_hidden = flm.get('num_hidden_sheets', 0)
        num_formula_cells = sum(s.get('cell_stats', {}).get('formula_cells', 0) for s in sheets)
        avg_text_len = 0
        textlens = [s.get('cell_stats', {}).get('avg_text_length', 0) for s in sheets if s.get('cell_stats')]
        if textlens:
            avg_text_len = sum(textlens) / len(textlens)

        num_vba = 1 if result.get('macros_and_scripts', {}).get('has_vba') else 0
        num_ext_links = len(result.get('external_references', {}).get('external_links', [])) if result.get('external_references') else 0
        num_embedded = len(result.get('embedded_items', {}).get('excel_embedded_files', [])) if result.get('embedded_items') else 0
        linked_domains = result.get('external_references', {}).get('linked_domains', []) if result.get('external_references') else []
        pii = result.get('text_extraction', {}).get('pii_counts', {})
        pii_total = sum(v for v in pii.values()) if isinstance(pii, dict) else 0

        features = {
            'num_sheets': int(num_sheets),
            'num_hidden_sheets': int(num_hidden),
            'num_formula_cells': int(num_formula_cells),
            'formula_complexity_mean': 0.0,  # placeholder (would require parsing every formula deeply)
            'num_vba_modules': int(num_vba),
            'num_external_links': int(num_ext_links),
            'num_embedded_files': int(num_embedded),
            'avg_text_len': float(avg_text_len),
            'entropy_of_binary': 0.0,
            'num_unique_domains': len(set(linked_domains)),
            'pii_count_total': int(pii_total),
        }
        result['structural_features_for_ml'] = {'features_vector': features}


